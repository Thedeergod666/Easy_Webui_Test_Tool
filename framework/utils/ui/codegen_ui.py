#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Codegen2Excel工具UI界面
此脚本提供一个简单的命令行交互界面，用于选择不同的操作模式
"""

import os
import sys
import argparse
from pathlib import Path

# 添加项目根目录到sys.path
project_root = Path(__file__).parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

# 导入codegen_to_excel模块中的函数
sys.path.insert(0, str(project_root / "framework" / "utils" / "codegen_to_excel"))

def check_flow_name_exists(flow_name):
    """检查测试流程名称是否已存在"""
    excel_file = project_root / "test_data" / f"{flow_name}.xlsx"
    return excel_file.exists()

def get_next_sheet_name(excel_file, base_sheet_name):
    """获取下一个可用的Sheet名称"""
    if not excel_file.exists():
        return base_sheet_name
    
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_file)
        existing_sheets = wb.sheetnames
        wb.close()
        
        # 根据用户的简化逻辑：如果有n个sheet，新生成的就叫Sheet(n+1)
        # 不管sheet名称是什么，只根据sheet的总数来确定下一个名称
        next_number = len(existing_sheets) + 1
        return f"Sheet{next_number}"
    except Exception:
        # 如果出现任何错误，返回基础名称
        return base_sheet_name

def convert_from_file():
    """从现有Python文件转换"""
    print("\n--- 从现有Python文件转换 ---")
    
    # 获取输入文件路径
    py_file = input("请输入Python文件路径: ").strip()
    if not py_file:
        print("[错误] 文件路径不能为空")
        input("按回车键返回主菜单...")
        return
    
    # 检查文件是否存在
    if not Path(py_file).exists():
        print(f"[错误] 文件不存在: {py_file}")
        input("按回车键返回主菜单...")
        return
    
    # 获取流程名称
    flow_name = input("请输入新测试流程的名称: ").strip()
    if not flow_name:
        print("[错误] 流程名称不能为空")
        input("按回车键返回主菜单...")
        return
    
    # 检查测试流程名称是否已存在
    if check_flow_name_exists(flow_name):
        print(f"[提示] 测试流程 '{flow_name}' 已存在")
    
    # 获取其他参数
    sheet_name_input = input("请输入Sheet名称 (默认: Sheet1): ").strip()
    sheet_name = sheet_name_input or "Sheet1"
    
    # 如果没有输入Sheet名称且文件已存在，自动获取下一个可用的Sheet名称
    if not sheet_name_input and check_flow_name_exists(flow_name):
        excel_file = project_root / "test_data" / f"{flow_name}.xlsx"
        sheet_name = get_next_sheet_name(excel_file, "Sheet1")
        print(f"[提示] 自动使用Sheet名称: {sheet_name}")
    
    browser = input("请输入浏览器类型 (默认: chromium, 可用简写: cr/chromium, ff/firefox, wk/webkit): ").strip() or "chromium"
    enabled = input("是否启用该流程? (y/n, 默认: y): ").strip().lower() != "n"
    
    # 构建命令
    cmd = [
        sys.executable, "-m", "framework.utils.codegen_to_excel.codegen_to_excel",
        py_file, flow_name,
        "--sheet-name", sheet_name,
        "--browser", browser
    ]
    
    if not enabled:
        cmd.append("--disabled")
    
    # 执行转换
    print(f"\n开始转换...")
    os.system(" ".join(cmd))
    input("\n按回车键返回主菜单...")

def record_and_convert():
    """启动Playwright录制并转换"""
    print("\n--- 启动Playwright录制并转换 ---")
    
    # 获取流程名称
    flow_name = input("请输入新测试流程的名称: ").strip()
    if not flow_name:
        print("[错误] 流程名称不能为空")
        input("按回车键返回主菜单...")
        return
    
    # 检查测试流程名称是否已存在
    if check_flow_name_exists(flow_name):
        print(f"[提示] 测试流程 '{flow_name}' 已存在")
    
    # 获取其他参数
    sheet_name_input = input("请输入Sheet名称 (默认: Sheet1): ").strip()
    sheet_name = sheet_name_input or "Sheet1"
    
    # 如果没有输入Sheet名称且文件已存在，自动获取下一个可用的Sheet名称
    if not sheet_name_input and check_flow_name_exists(flow_name):
        excel_file = project_root / "test_data" / f"{flow_name}.xlsx"
        sheet_name = get_next_sheet_name(excel_file, "Sheet1")
        print(f"[提示] 自动使用Sheet名称: {sheet_name}")
    
    browser = input("请输入浏览器类型 (默认: chromium, 可用简写: cr/chromium, ff/firefox, wk/webkit): ").strip() or "chromium"
    enabled = input("是否启用该流程? (y/n, 默认: y): ").strip().lower() != "n"
    
    # 构建命令
    cmd = [
        sys.executable, "-m", "framework.utils.codegen_to_excel.record_and_convert",
        flow_name,
        "--sheet-name", sheet_name,
        "--browser", browser
    ]
    
    if not enabled:
        cmd.append("--disabled")
    
    # 执行录制和转换
    print(f"\n开始录制和转换...")
    os.system(" ".join(cmd))
    input("\n按回车键返回主菜单...")

def show_menu():
    """显示Codegen工具菜单"""
    while True:
        os.system('cls' if os.name == 'nt' else 'clear')  # 清屏
        print("=" * 50)
        print("       Codegen2Excel工具")
        print("=" * 50)
        print()
        print("请选择操作:")
        print("  1. 从现有Python文件转换")
        print("  2. 启动Playwright录制并转换")
        print("  3. 返回主菜单")
        print()
        choice = input("请输入您的选择 [1, 2, 3]: ").strip()
        
        if choice == "1":
            convert_from_file()
        elif choice == "2":
            record_and_convert()
        elif choice == "3":
            return  # 返回主菜单
        else:
            print("无效输入，请重试...")
            input("按回车键继续...")

def run_with_menu():
    """运行Codegen工具菜单"""
    show_menu()

def run_with_args(convert=False, record=False):
    """根据命令行参数运行Codegen工具"""
    if convert:
        convert_from_file()
    elif record:
        record_and_convert()