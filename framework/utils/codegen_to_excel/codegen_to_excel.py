# framework/utils/codegen_to_excel.py (V3 - 终极修复版)
import ast
import json
import os
import sys
import pandas as pd
from argparse import ArgumentParser

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

ACTION_MAP = {
    'goto': 'open',
    'click': 'click',
    'fill': 'on_input',
    'press': 'press',
    'set_input_files': 'upload_file',
}

class CodeGenParser(ast.NodeVisitor):
    def __init__(self, total_nodes=0):
        self.steps = []
        self.page_vars = {'page', 'page1', 'page2', 'page3', 'page4', 'page5', 'page6'}
        self.page_var_mapping = {'page': 1}  # 页面变量到页面索引的映射，默认page为第1页
        self.current_page = 1  # 当前活动页面索引
        self.next_page_index = 2  # 下一个页面的索引
        self.processed_nodes = 0
        self.total_nodes = total_nodes
        self.next_case_number = 0  # 下一个可用的case编号

    def update_progress(self):
        """更新进度条显示"""
        self.processed_nodes += 1
        if self.total_nodes > 0:
            percentage = (self.processed_nodes / self.total_nodes) * 100
            print(f"\r处理进度: {self.processed_nodes}/{self.total_nodes} ({percentage:.1f}%)", end='', flush=True)

    def visit_Expr(self, node):
        if isinstance(node.value, ast.Call):
            self.parse_action_call(node.value)
        self.update_progress()  # 更新进度
        self.generic_visit(node)
    
    def visit_With(self, node):
        """处理with语句，识别page.expect_popup()模式"""
        # 检查是否是page.expect_popup()模式
        if (len(node.items) == 1 and
            isinstance(node.items[0].context_expr, ast.Call) and
            isinstance(node.items[0].context_expr.func, ast.Attribute) and
            node.items[0].context_expr.func.attr == 'expect_popup'):
            
            # 获取页面变量名，如page1_info
            page_info_var = None
            if isinstance(node.items[0].optional_vars, ast.Name):
                page_info_var = node.items[0].optional_vars.id
            
            # 处理with块内的语句
            for item in node.body:
                if isinstance(item, ast.Expr) and isinstance(item.value, ast.Call):
                    self.parse_action_call(item.value)
                    
                    # 如果是点击操作，生成页面切换步骤
                    action_name = self.get_final_action_name(item.value)
                    if action_name == 'click':
                        # 创建新的页面变量
                        if page_info_var:
                            # 从page1_info推断出page1
                            page_var = page_info_var.replace('_info', '')
                            if page_var not in self.page_var_mapping:
                                self.page_var_mapping[page_var] = self.next_page_index
                                self.next_page_index += 1
                            
                            # 生成切换到新页面的步骤
                            switch_row = self.create_base_row('switch_to_page', 'switch_to_page')
                            switch_row['数据内容'] = str(self.page_var_mapping[page_var])
                            switch_row['描述'] = f'切换到新页面 {page_var}'
                            self.steps.append(switch_row)
                            
                            # 更新当前活动页面
                            self.current_page = self.page_var_mapping[page_var]
        else:
            # 处理其他with语句
            self.generic_visit(node)
    
    def visit_Assign(self, node):
        """处理赋值语句，识别page1 = page1_info.value这样的语句"""
        # 检查是否是page1 = page1_info.value模式
        if (len(node.targets) == 1 and
            isinstance(node.targets[0], ast.Name) and
            isinstance(node.value, ast.Attribute) and
            node.value.attr == 'value'):
            
            # 获取目标变量名，如page1
            target_var = node.targets[0].id
            
            # 获取源变量名，如page1_info
            if isinstance(node.value.value, ast.Name):
                source_var = node.value.value.id
                
                # 检查是否是page_info模式
                if source_var.endswith('_info'):
                    # 从page1_info推断出page1
                    page_var = source_var.replace('_info', '')
                    if page_var in self.page_vars and page_var not in self.page_var_mapping:
                        self.page_var_mapping[page_var] = self.next_page_index
                        self.next_page_index += 1
                        
                        # 同时更新目标变量的映射
                        if target_var != page_var:
                            self.page_var_mapping[target_var] = self.page_var_mapping[page_var]
        
        # 处理其他赋值语句
        self.generic_visit(node)

    # ▼▼▼【核心修复】▼▼▼
    def is_expect_chain(self, call_node) -> bool:
        """
        递归地向上追溯调用链，判断其根源是否为 'expect' 函数。
        这是正确处理 expect(...).to_be_visible() 等链式断言的关键。
        """
        current_node = call_node
        # 循环向上查找，直到找到链的起点或非调用节点
        while isinstance(current_node, ast.Call):
            func_node = current_node.func
            if isinstance(func_node, ast.Name) and func_node.id == 'expect':
                return True # 找到了！
            if isinstance(func_node, ast.Attribute):
                current_node = func_node.value
            else:
                return False # 链中断了，且不是expect
        return False
    
    def parse_action_call(self, call_node):
        original_code = self.unparse_node(call_node, clean_quotes=False)

        # 1. 使用新的、更强大的 `is_expect_chain` 函数来做判断
        if self.is_expect_chain(call_node):
            self.handle_expect_call(call_node, original_code)
            return

        # 2. 检查是否是页面跳转语句，如 page1.goto(...)
        # 这类语句需要放在相应的断言之前
        if (isinstance(call_node, ast.Call) and
            isinstance(call_node.func, ast.Attribute) and
            call_node.func.attr == 'goto' and
            isinstance(call_node.func.value, ast.Name) and
            call_node.func.value.id in self.page_vars):
            
            # 创建页面跳转步骤
            page_var = call_node.func.value.id
            row = self.create_base_row('open', 'goto')
            row['数据内容'] = self.unparse_node(call_node.args[0])
            row['补充说明'] = f"原始代码: {original_code}"
            
            # 查找相应的断言步骤，并将页面跳转步骤插入到其之前
            # 通过查找包含相同页面变量名的断言步骤
            inserted = False
            for i, step in enumerate(self.steps):
                if (step['关键字'] == 'expect_codegen' and
                    page_var in step['目标对象']):
                    self.steps.insert(i, row)
                    inserted = True
                    break
            
            # 如果没有找到相应的断言步骤，则添加到步骤列表末尾
            if not inserted:
                self.steps.append(row)
            return

        # 3. 后续逻辑保持不变
        action_name = self.get_final_action_name(call_node)
        keyword = ACTION_MAP.get(action_name)

        if not keyword:
            print(f"[信息] 暂不支持的动作: '{action_name}'，将标记为'skip'。")
            row = self.create_base_row('', action_name)
            row['执行状态'] = 'skip'
            row['补充说明'] = f"不支持的动作, 原始代码: {original_code}"
            self.steps.append(row)
            return

        row = self.create_base_row(keyword, action_name)
        row['补充说明'] = f"原始代码: {original_code}"
        
        if keyword == 'open':
            row['数据内容'] = self.unparse_node(call_node.args[0])
        else:
            locator_str, data_content = self.extract_locator_and_data(call_node, action_name)
            row['定位方式'] = 'codegen'
            row['目标对象'] = locator_str
            if data_content:
                row['数据内容'] = data_content

        self.steps.append(row)

    def handle_expect_call(self, call_node, original_code):
        row = self.create_base_row('expect_codegen', '断言')
        row['目标对象'] = original_code
        row['补充说明'] = f"原始代码: {original_code}"
        self.steps.append(row)

    def get_final_action_name(self, node):
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Attribute):
            return node.func.attr
        return None

    def unparse_node(self, node, clean_quotes=True):
        code_str = ast.unparse(node).strip()
        if clean_quotes and ((code_str.startswith("'") and code_str.endswith("'")) or \
                             (code_str.startswith('"') and code_str.endswith('"'))):
            return code_str[1:-1]
        return code_str

    def extract_locator_and_data(self, call_node, action):
        data_content = ""
        if call_node.args:
            data_content = self.unparse_node(call_node.args[0])
        
        locator_node = call_node.func.value
        locator_full_str = self.unparse_node(locator_node, clean_quotes=False)
        
        for page_var in self.page_vars:
            if locator_full_str.startswith(f"{page_var}."):
                # 检查是否需要页面切换
                page_index = self.page_var_mapping.get(page_var, 1)
                if page_index != self.current_page:
                    # 生成页面切换步骤
                    switch_row = self.create_base_row('switch_to_page', 'switch_to_page')
                    switch_row['数据内容'] = str(page_index)
                    switch_row['描述'] = f'切换到页面 {page_var}'
                    self.steps.append(switch_row)
                    self.current_page = page_index
                
                return locator_full_str[len(page_var)+1:], data_content
        return locator_full_str, data_content

    def create_base_row(self, keyword, action_name=''):
        case_number = self.next_case_number
        self.next_case_number += 1
        return {
            '编号': f'case_{case_number:03d}',
            '关键字': keyword, '验证类型': '', '定位方式': '', '目标对象': '',
            '数据内容': '', '描述': f'自动生成: {action_name}', '执行状态': '', '补充说明': ''
        }
        
def apply_excel_styles(workbook, worksheet, df):
    """应用Excel样式"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # 定义样式
    header_font = Font(name='微软雅黑', size=14, bold=True)
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    normal_font = Font(name='微软雅黑', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)  # 首行不自动换行
    normal_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    
    # 定义边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 设置第一行样式
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment  # 首行不自动换行
        cell.border = thin_border
    
    # 设置其他行样式
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = normal_font
            cell.alignment = normal_alignment
            cell.border = thin_border
            
            # 关键字列内容为open时填充绿色
            if col == 2 and cell.value == 'open':  # 第2列是'关键字'列
                cell.fill = green_fill
    
    # 设置行高
    worksheet.row_dimensions[1].height = 34.45  # 第一行行高34.45磅
    for row in range(2, len(df) + 2):
        worksheet.row_dimensions[row].height = 28.8  # 其他行行高28.8磅
    
    # 自动调整列宽，最宽不超过56字符，最窄不低于完整展现首行文字所需的宽度
    for col in range(1, len(df.columns) + 1):
        column = get_column_letter(col)
        
        # 计算列标题的长度
        header_length = len(str(df.columns[col-1]))
        
        # 计算列数据的最大长度
        max_data_length = 0
        for cell in worksheet[column][1:]:  # 从第二行开始计算数据长度
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_data_length:
                    max_data_length = cell_length
            except:
                pass
        
        # 取列标题和列数据的最大长度
        max_length = max(header_length, max_data_length)
        
        # 限制宽度在合理范围内：最小为列标题长度，最大为56字符
        adjusted_width = max(header_length, min(max_length + 2, 56))
        worksheet.column_dimensions[column].width = adjusted_width
    
    # 首行冻结
    worksheet.freeze_panes = 'A2'
    
    return workbook, worksheet

def convert_py_to_excel(py_file, output_excel, sheet_name='Sheet1'):
    """主转换函数"""
    with open(py_file, 'r', encoding='utf-8') as f:
        source_code = f.read()

    tree = ast.parse(source_code)
    
    # 计算总的表达式节点数用于进度条
    total_nodes = sum(1 for node in ast.walk(tree) if isinstance(node, ast.Expr))
    
    parser = CodeGenParser(total_nodes)
    parser.visit(tree)
    
    # 完成后换行
    if total_nodes > 0:
        print()  # 换行

    if not parser.steps:
        print("[错误] 未能从Python文件中解析出任何测试步骤。")
        return False
    
    # 重新为所有步骤分配编号，确保编号是连续且正确的
    for i, step in enumerate(parser.steps):
        step['编号'] = f'case_{i:03d}'
    
    df = pd.DataFrame(parser.steps)
    df = df[['编号', '关键字', '验证类型', '定位方式', '目标对象', '数据内容', '描述', '执行状态', '补充说明']]
    
    # 检查Excel文件是否已存在
    if os.path.exists(output_excel):
        # 如果文件已存在，检查Sheet名称是否重复
        from openpyxl import load_workbook
        wb = load_workbook(output_excel)
        existing_sheets = wb.sheetnames
        
        # 如果Sheet名称重复，添加后缀
        final_sheet_name = sheet_name
        counter = 1
        while final_sheet_name in existing_sheets:
            final_sheet_name = f"{sheet_name}_{counter}"
            counter += 1
        
        # 添加新的sheet
        try:
            with pd.ExcelWriter(output_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, index=False, sheet_name=final_sheet_name)
            
            # 应用样式（在pd.ExcelWriter上下文管理器外部）
            wb = load_workbook(output_excel)
            ws = wb[final_sheet_name]
            apply_excel_styles(wb, ws, df)
            wb.save(output_excel)
        except PermissionError:
            print(f"[错误] 无法写入文件 '{output_excel}'，可能是文件正在被其他程序使用。请关闭文件后重试。")
            return False, final_sheet_name
        except Exception as e:
            print(f"[错误] 写入文件时出现异常: {e}")
            return False, final_sheet_name
    else:
        # 如果文件不存在，创建新文件
        df.to_excel(output_excel, index=False, sheet_name=sheet_name)
        final_sheet_name = sheet_name
        
        # 应用样式
        from openpyxl import load_workbook
        wb = load_workbook(output_excel)
        ws = wb[sheet_name]
        apply_excel_styles(wb, ws, df)
        wb.save(output_excel)
    
    return True, final_sheet_name

def update_test_config(output_excel_path, flow_name, sheet_name="Sheet1", browser="chromium", enabled=True):
    """更新test_config.json文件，添加新生成的测试流程"""
    config_path = os.path.join(project_root, 'test_data', 'test_config.json')
    
    # 读取现有的配置文件
    if os.path.exists(config_path):
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
    else:
        # 如果配置文件不存在，创建一个默认的配置
        config = {
            "visual_mode": {
                "headed": True,
                "slow_mo": 50
            },
            "test_flows": []
        }
    
    # 创建相对路径
    relative_path = os.path.relpath(output_excel_path, project_root).replace("\\", "/")
    
    # 创建新的测试流程配置
    new_flow = {
        "file_path": relative_path,
        "sheet_name": sheet_name,
        "description": f"自动生成的测试流程: {flow_name}",
        "browser": browser,
        "enabled": enabled
    }
    
    # 检查是否已存在相同的配置
    existing_flows = config.get("test_flows", [])
    flow_exists = False
    for i, flow in enumerate(existing_flows):
        if flow.get("file_path") == relative_path and flow.get("sheet_name") == sheet_name:
            # 更新现有的配置
            existing_flows[i] = new_flow
            flow_exists = True
            print(f"  > 已更新现有的测试流程配置")
            break
    
    # 如果不存在，则添加新的配置
    if not flow_exists:
        config["test_flows"].append(new_flow)
        print(f"  > 已添加新的测试流程配置")
    
    # 保存更新后的配置文件
    with open(config_path, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=4)
    
    return True

if __name__ == '__main__':
    parser = ArgumentParser(description="将Playwright Codegen生成的PY文件转换为关键字驱动的Excel文件。")
    parser.add_argument("py_file", help="输入的.py文件路径")
    parser.add_argument("flow_name", help="新测试流程的名称 (将作为Excel文件名)")
    parser.add_argument("--sheet-name", default="Sheet1", help="指定Excel中的sheet名称 (默认: Sheet1)")
    parser.add_argument("--no-config-update", action="store_true", help="不自动更新test_config.json配置文件")
    parser.add_argument("--browser", default="chromium", help="指定测试浏览器 (默认: chromium)")
    parser.add_argument("--disabled", action="store_true", help="将新流程设置为禁用状态")
    args = parser.parse_args()

    output_excel_path = os.path.join(project_root, 'test_data', f"{args.flow_name}.xlsx")
    
    print(f"--- Codegen to Excel 转换器 V3 ---")
    print(f"  > 输入文件: {args.py_file}")
    
    result = convert_py_to_excel(args.py_file, output_excel_path, args.sheet_name)
    if result[0]:  # 检查成功状态
        final_sheet_name = result[1]  # 获取最终的Sheet名称
        print(f"  > Excel 文件已成功生成: {output_excel_path}")
        print(f"  > Sheet 名称: {final_sheet_name}")
        
        # 如果没有指定--no-config-update参数，则自动更新配置文件
        if not args.no_config_update:
            if update_test_config(output_excel_path, args.flow_name, final_sheet_name, args.browser, not args.disabled):
                print(f"  > test_config.json 配置文件已更新")
            else:
                print(f"  > test_config.json 配置文件更新失败")
    else:
        print(f"--- 转换失败 ---")
        final_sheet_name = result[1]  # 获取最终的Sheet名称
        if final_sheet_name:  # 如果有Sheet名称，也显示出来
            print(f"  > Sheet 名称: {final_sheet_name}")
